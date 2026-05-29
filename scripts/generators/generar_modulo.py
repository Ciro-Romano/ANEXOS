from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl.worksheet.table import Table
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils import get_column_letter
import shutil
import os


TEMPLATE_PATH = "templates/Modulo-00.xlsx"
EXPORT_PATH = "scripts/exports"

def formatear_cuit(cuit):

    cuit = cuit.strip()

    cuit = "".join(filter(str.isdigit, cuit))

    if len(cuit) != 11:
        raise ValueError("CUIT Invalido")
    
    return f"{cuit[:2]}-{cuit[2:10]}-{cuit[10]}"

def generar_periodos(desde, hasta):

    fecha_desde = datetime.strptime(
        desde,
        "%Y-%m"
    )

    fecha_hasta = datetime.strptime(
        hasta,
        "%Y-%m"
    )

    periodos = []

    while fecha_desde <= fecha_hasta:

        periodos.append(
            fecha_desde.strftime("%Y-%m")
        )

        fecha_desde += relativedelta(
            months=1
        )

    return periodos


def generar_modulo(empresa,cuit,desde,hasta,acta):

    cuit_formateado = formatear_cuit(cuit)

    caracteres_invalidos = ["/", ":", "|"]

    for caracter in caracteres_invalidos:
       empresa = empresa.replace(caracter, "-")

    empresa_nombre = empresa.upper()

    nombre_archivo = f"Acta N° {acta} - {empresa_nombre}.xlsx"
    ruta_salida = os.path.join(EXPORT_PATH,nombre_archivo)

    # copia de template
    shutil.copy(TEMPLATE_PATH, ruta_salida)

    # abrir excel
    wb = load_workbook(ruta_salida)

    periodos = generar_periodos(
        desde,
        hasta
    )
    
    # HOJA BASE
    ws_base = wb.active
    ws_base.title = "_BASE"
    ws_base.sheet_state = "hidden"

    tabla_original = list(ws_base.tables.values())[0]

    wb_modulo = load_workbook(
        "templates/Modulo-02-04-00.xlsx"
    )

    ws_modulo = wb_modulo.active

    tabla_modulo = list(
        ws_modulo.tables.values()
    )[0]

    min_col_core, min_row_core, max_col_core, max_row_core = (
        range_boundaries(tabla_original.ref)
    )

    min_col_mod, min_row_mod, max_col_mod, max_row_mod = (
        range_boundaries(tabla_modulo.ref)
    )

    columna_destino = max_col_core + 1

    for col in range(max_col_core + 1, max_col_mod + 1):

        encabezado = ws_modulo.cell(
            min_row_mod,
            col
        ).value

        ws_base.cell(
            min_row_core,
            columna_destino
        ).value = encabezado

        columna_destino += 1

    nuevo_max_col = max_col_mod

    tabla_original.ref = (
        f"{get_column_letter(min_col_core)}{min_row_core}:"
        f"{get_column_letter(nuevo_max_col)}{max_row_core}"
    )

    print("nuevo ref =", tabla_original.ref)

    # CREAR HOJAS
    for periodo in periodos:

        ws = wb.copy_worksheet(ws_base)

        ws.title = periodo

        for table_name in list(ws.tables.keys()):
            del ws.tables[table_name]

        periodo_tabla = periodo.replace("-", "")[2:]

        nueva_tabla = Table(
            displayName=f"Modulo_00_{periodo_tabla}",
            ref=tabla_original.ref
        )

        nueva_tabla.tableStyleInfo = tabla_original.tableStyleInfo

        nueva_tabla.showAutoFilter = False

        ws.add_table(nueva_tabla)

    for col in range(max_col_core + 1, max_col_mod + 1):

        letra = get_column_letter(col)

        ws_base.column_dimensions[
            letra
        ].width = ws_modulo.column_dimensions[
            letra
        ].width

        # DATOS
        ws["B4"] = empresa.upper()
        ws["B4"].font = Font(
            name="Arial",   #tipo de letra
            size=20,        #tamaño        
            bold=True,      #negrita
            italic=True,    #cursiva
        )
        ws["B4"].alignment = Alignment(
            vertical="bottom",
            horizontal="left",
        )

        ws["C5"] = cuit_formateado
        ws["C5"].font = Font(
            name="Arial",
            size=11,
            italic=True,
        )
        ws["C5"].alignment = Alignment(
            vertical="bottom",
            horizontal="right",
        )

        ws["I6"] = periodo
        ws["I6"].font = Font(
            name="Arial",
            size=11,
            italic=True
        )
        ws["I6"].alignment = Alignment(
            vertical="bottom",
            horizontal="right",
        )

        ws["C6"] = acta
        ws["C6"].font = Font(
            name="Arial",
            size=11,
            italic=True,
            )
        ws["C6"].alignment = Alignment(
            vertical="bottom",
            horizontal="right",
        )
    
    wb.save(ruta_salida)

    return ruta_salida

    